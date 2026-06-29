from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "BRD Use Case Status"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="14384D", end_color="14384D", fill_type="solid")
category_font = Font(bold=True, size=11, color="14384D")
category_fill = PatternFill(start_color="E8ECEF", end_color="E8ECEF", fill_type="solid")
green_font = Font(color="006100", bold=True)
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_font = Font(color="9C6500", bold=True)
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_font = Font(color="9C0006", bold=True)
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

headers = [
    "UC ID", "Use Case Name", "Module", "Priority",
    "Backend Status", "Frontend Status", "Overall Status", "Notes / Gaps"
]

col_widths = [10, 35, 25, 10, 18, 18, 18, 55]
for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(i)].width = w

# Data: (UC_ID, Name, Module, Priority, BE, FE, Overall, Notes)
data = [
    # --- AUTH MODULE ---
    ("", "MODULE: USER AUTHENTICATION & REGISTRATION", "", "", "", "", "", ""),
    ("UC-01", "Register New Student Account",
     "Auth", "Critical",
     "✅", "✅", "✅",
     "BE: Migration 0007 adds DOB/country/language/is_verified fields + email_verification_tokens. "
     "Register returns JWT tokens + verification token logged in dev. verify-email + resend endpoints exist. "
     "FE: LoginPage handles register mode, authStore manages tokens + user. "
     "Gap: No actual SMTP email sending (tokens logged to console only in dev)."),
    ("UC-02", "Authenticate User (Login)",
     "Auth", "Critical",
     "✅", "✅", "✅",
     "BE: JWT access token (15min) + refresh token (7d). bcrypt password verification. "
     "FE: LoginPage + authStore.login() + automatic token refresh on 401."),
    ("UC-03", "Reset Forgotten Password",
     "Auth", "High",
     "✅", "⚠️", "⚠️",
     "BE: forgot-password + reset-password endpoints exist. Tokens expire in 1hr. "
     "FE: 'Forgot password?' link on LoginPage navigates to /login?forgot=true — no actual forgot/reset page implemented. "
     "Gap: No UI for password reset flow."),
    ("UC-04", "Link Parent to Student Account",
     "Auth", "High",
     "⚠️", "⚠️", "⚠️",
     "BE: ParentStudentLink model exists. Parent dashboard endpoints retrieve student data. "
     "FE: ParentDashboard lists students. "
     "Gap: No explicit linking-code generation/validation flow. No dedicated link/unlink endpoint. "
     "Parent sees first child only."),

    # --- STUDENT DASHBOARD ---
    ("", "MODULE: STUDENT DASHBOARD", "", "", "", "", "", ""),
    ("UC-05", "View Personalized Student Dashboard",
     "Student", "Critical",
     "✅", "✅", "✅",
     "BE: Enrollment + mastery + curriculum endpoints provide dashboard data. "
     "FE: StudentDashboard page with stats, learning path mini-view, quick actions. "
     "Gap: No streak counter or achievement badges displayed (gamification data not tracked)."),

    # --- COURSE DISCOVERY ---
    ("", "MODULE: COURSE DISCOVERY & ENROLLMENT", "", "", "", "", "", ""),
    ("UC-06", "Search & Browse Course Catalog",
     "Discovery", "High",
     "❌", "❌", "❌",
     "BE: No search/catalog endpoint exists. getCourses() has search param but no full-text search. "
     "FE: No course catalog page built. "
     "Gap: Major missing feature — no course browsing UI at all."),
    ("UC-07", "Enroll in a Course",
     "Enrollment", "Critical",
     "✅", "✅", "✅",
     "BE: Enrollment router (POST /enrollments, GET /enrollments). "
     "FE: StudentOnboarding auto-enrolls into first published course. Enrollment list on dashboard. "
     "Gap: No prerequisite validation before enrollment (student can enroll without meeting prerequisites)."),

    # --- CURRICULUM ---
    ("", "MODULE: COURSE & CURRICULUM MANAGEMENT", "", "", "", "", "", ""),
    ("UC-08", "Create New Course",
     "Curriculum", "Critical",
     "✅", "⚠️", "⚠️",
     "BE: Full CRUD for courses. "
     "FE: Admin/Teacher course creation form NOT built (only AssessmentManagement page exists for creation). "
     "Gap: No course creation UI for teachers."),
    ("UC-09", "Create Curriculum Hierarchy (Modules->Lessons->Concepts)",
     "Curriculum", "Critical",
     "✅", "⚠️", "⚠️",
     "BE: Full CRUD for modules, lessons, concepts, contents, exercises, examples, objectives. "
     "FE: Curriculum tree fetched and displayed (LearningPathPage, ProgressPage). "
     "Gap: No curriculum builder UI for teachers to create/edit the hierarchy."),
    ("UC-10", "Publish Course for Enrollment",
     "Curriculum", "High",
     "✅", "❌", "⚠️",
     "BE: Publish/unpublish endpoints exist for courses. "
     "FE: No UI for publishing courses. "
     "Gap: No publish button/flow in frontend."),

    # --- AI TUTOR ---
    ("", "MODULE: AI TUTOR", "", "", "", "", "", ""),
    ("UC-11", "Start AI Tutor Session",
     "AI Tutor", "Critical",
     "✅", "✅", "✅",
     "BE: Teaching session endpoints (create, get, pause, end). LangGraph pipeline. "
     "FE: AITutorPage auto-starts session + auto-ends on unmount."),
    ("UC-12", "AI Adaptive Teaching Based on Student Response",
     "AI Tutor", "Critical",
     "✅", "✅", "✅",
     "BE: 7-node LangGraph pipeline (retrieve_memories -> teach -> ask_question -> evaluate_response -> diagnose -> provide_example/complete_concept). Gemini 1.5 Pro with mock fallback. "
     "FE: Chat interface sends student input via teach() endpoint."),
    ("UC-13", "End AI Tutor Session with Summary",
     "AI Tutor", "Critical",
     "✅", "⚠️", "⚠️",
     "BE: Session end/pause endpoints. "
     "FE: Session ends on unmount. "
     "Gap: No end-of-session summary displayed to student (BE likely returns data but FE doesn't show it)."),

    # --- ASSESSMENT ---
    ("", "MODULE: ASSESSMENT & MASTERY", "", "", "", "", "", ""),
    ("UC-14", "Take Lesson Quiz / Assessment",
     "Assessment", "Critical",
     "✅", "✅", "✅",
     "BE: Full assessment CRUD, questions, attempts, auto-grading. "
     "FE: Full attempt flow (list -> detail -> attempt with timer -> submit -> results with review)."),
    ("UC-15", "View Concept-by-Concept Mastery Progress",
     "Assessment", "High",
     "✅", "✅", "✅",
     "BE: Mastery router (overview, by concept, by course). "
     "FE: ProgressPage with mastery bar chart, areas to improve, strong areas."),

    # --- KNOWLEDGE GRAPH ---
    ("", "MODULE: KNOWLEDGE GRAPH & LEARNING PATH", "", "", "", "", "", ""),
    ("UC-16", "Maintain Knowledge Graph (Concepts & Prerequisites)",
     "Knowledge Graph", "High",
     "⚠️", "❌", "⚠️",
     "BE: Knowledge graph router (create edge, delete edge, get prerequisites). "
     "Gap: No UI for teachers to manage the knowledge graph. Admin-only access currently."),
    ("UC-17", "Generate Adaptive Learning Path",
     "Knowledge Graph", "High",
     "✅", "✅", "✅",
     "BE: Prerequisites from knowledge graph inform curriculum ordering. "
     "FE: LearningPathPage displays ordered lesson list with status."),

    # --- TEACHER ---
    ("", "MODULE: TEACHER DASHBOARD", "", "", "", "", "", ""),
    ("UC-18", "View Course Analytics (Teacher)",
     "Teacher", "High",
     "✅", "✅", "✅",
     "BE: Teacher dashboard endpoints (students, progress, mastery, misconceptions, sessions, attempts, courses). "
     "FE: TeacherDashboard with stat cards + students list + recent sessions."),
    ("UC-19", "Identify & Support At-Risk Students",
     "Teacher", "Critical",
     "⚠️", "⚠️", "⚠️",
     "BE: Teacher can access individual student data (progress, mastery, misconceptions). "
     "FE: TeacherStudentDetail shows full student profile. "
     "Gap: No automated at-risk flagging (declining scores, engagement drops). No intervention tools (video upload, message). "
     "Teacher must manually identify struggling students."),
    ("UC-20", "Create & Monetize Courses",
     "Teacher", "High",
     "❌", "❌", "❌",
     "BE: No payment system, no revenue sharing, no payout processing. "
     "FE: No monetization UI. "
     "Gap: Entire payments/monetization module missing."),

    # --- PARENT ---
    ("", "MODULE: PARENT DASHBOARD", "", "", "", "", "", ""),
    ("UC-21", "View Child's Detailed Progress Report",
     "Parent", "Critical",
     "✅", "✅", "✅",
     "BE: 11 parent dashboard endpoints (students, progress, curriculum, mastery, misconceptions, sessions, activity, dashboard, reports). "
     "FE: ParentDashboard + ParentStudentDetail + ReportsPage + InsightsPage."),
    ("UC-22", "Configure Parental Controls (Time Limits, Sleep Mode)",
     "Parent", "High",
     "❌", "❌", "❌",
     "BE: No usage limit, break reminder, or sleep mode endpoints. "
     "FE: No parental controls UI. "
     "Gap: Entire safety/controls module missing."),
    ("UC-23", "Communicate with Teacher",
     "Parent", "Medium",
     "❌", "❌", "❌",
     "BE: No messaging/communication endpoints. "
     "FE: No messaging UI. "
     "Gap: Entire messaging module missing."),

    # --- ADMIN ---
    ("", "MODULE: ADMINISTRATOR", "", "", "", "", "", ""),
    ("UC-24", "View Platform-Wide Analytics",
     "Admin", "High",
     "✅", "✅", "✅",
     "BE: Admin dashboard stats endpoint. Analytics overview + per-course analytics. "
     "FE: AdminDashboard + AnalyticsPage with system stats, per-course analysis."),
    ("UC-25", "Moderate User-Generated Content",
     "Admin", "High",
     "⚠️", "⚠️", "⚠️",
     "BE: Content ingestion draft/approve/publish flow exists. "
     "FE: AssessmentManagementPage for creating assessments. "
     "Gap: No content moderation queue UI. No community ratings. No automated quality checks."),

    # --- INSTITUTIONAL ---
    ("", "MODULE: INSTITUTIONAL / SCHOOL DISTRICT", "", "", "", "", "", ""),
    ("UC-26", "Onboard School District",
     "Institutional", "High",
     "❌", "❌", "❌",
     "Gap: Entire institutional module missing. No multi-tenant support, no district/school hierarchy, no bulk enrollment via CSV/SSO."),
    ("UC-27", "Generate Institutional Reports",
     "Institutional", "High",
     "❌", "❌", "❌",
     "Gap: No district-level analytics, no cross-school comparison reports, no auto-generated monthly reports for boards."),

    # --- GAMIFICATION ---
    ("", "MODULE: GAMIFICATION & ACHIEVEMENT", "", "", "", "", "", ""),
    ("UC-28", "Award Achievements & Badges",
     "Gamification", "Medium",
     "❌", "❌", "❌",
     "BE: current_streak_days on StudentProfile is the only gamification data point. "
     "Gap: No achievements system, no badges, no XP, no triggers."),
    ("UC-29", "Issue Certificate of Completion",
     "Gamification", "High",
     "❌", "❌", "❌",
     "Gap: No certificate generation (PDF), no verification URLs, no serial numbers."),

    # --- NOTIFICATIONS ---
    ("", "MODULE: NOTIFICATION & COMMUNICATION", "", "", "", "", "", ""),
    ("UC-30", "Send Progress Notification to Parent",
     "Notifications", "High",
     "❌", "❌", "❌",
     "BE: Email service logs to console only (stub). No real SMTP integration. No in-app notification system. "
     "FE: No notification UI. "
     "Gap: Entire notification engine missing in production."),

    # --- PAYMENTS ---
    ("", "MODULE: PAYMENT & MONETIZATION", "", "", "", "", "", ""),
    ("UC-31", "Purchase a Course",
     "Payments", "High",
     "❌", "❌", "❌",
     "Gap: Entire payment module missing (no Stripe/PayPal, no checkout flow, no receipts)."),
    ("UC-32", "Process Teacher Payouts",
     "Payments", "High",
     "❌", "❌", "❌",
     "Gap: No payout system, no earnings tracking, no revenue sharing."),

    # --- CONTENT INGESTION ---
    ("", "MODULE: CONTENT INGESTION & UPLOAD", "", "", "", "", "", ""),
    ("UC-33", "Upload & Parse Teaching Material (PDF/DOCX/TXT)",
     "Content", "Medium",
     "✅", "❌", "⚠️",
     "BE: Full content ingestion pipeline (upload, parse, draft CRUD, generate, approve, publish). Supports PDF/DOCX/TXT. "
     "FE: No upload UI built. "
     "Gap: Teachers cannot upload files via frontend."),

    # --- SAFETY ---
    ("", "MODULE: SAFETY, PRIVACY & PARENTAL CONTROLS", "", "", "", "", "", ""),
    ("UC-34", "Enforce Usage Limits & Break Reminders",
     "Safety", "Critical",
     "❌", "❌", "❌",
     "Gap: No usage limits, no break reminders, no sleep mode enforcement. Entire safety module missing."),
    ("UC-35", "Enforce Data Privacy (COPPA/GDPR)",
     "Safety", "Critical",
     "⚠️", "⚠️", "⚠️",
     "BE: JWT auth, bcrypt passwords, CORS, security headers. Audit models exist. "
     "FE: None. "
     "Gap: No COPPA consent flow, no data export (GDPR right to access), no account deletion endpoint, no audit log writing."),

    # --- OFFLINE ---
    ("", "MODULE: OFFLINE & ACCESSIBILITY", "", "", "", "", "", ""),
    ("UC-36", "Download Lessons for Offline Use",
     "Offline", "Medium",
     "❌", "❌", "❌",
     "Gap: Entire offline mode missing. No content packaging/download, no offline sync."),
]

def set_cell(row, col, value, font=None, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = wrap
    cell.border = thin_border
    if font:
        cell.font = font
    if fill:
        cell.fill = fill

for row_idx, row_data in enumerate(data, 2):
    uc_id, name, module, priority, be, fe, overall, notes = row_data

    # Section headers (empty UC_ID)
    if not uc_id:
        for col in range(1, 9):
            set_cell(row_idx, col, "", category_font, category_fill)
        set_cell(row_idx, 2, name, category_font, category_fill)
        ws.cell(row=row_idx, column=2).font = Font(bold=True, size=11, color="14384D")
        continue

    set_cell(row_idx, 1, uc_id)
    set_cell(row_idx, 2, name)
    set_cell(row_idx, 3, module)
    set_cell(row_idx, 4, priority)

    # Status color coding
    for col, val in [(5, be), (6, fe), (7, overall)]:
        if val.startswith("✅"):
            set_cell(row_idx, col, val, green_font, green_fill)
        elif val.startswith("⚠️"):
            set_cell(row_idx, col, val, yellow_font, yellow_fill)
        elif val.startswith("❌"):
            set_cell(row_idx, col, val, red_font, red_fill)
        else:
            set_cell(row_idx, col, val)

    set_cell(row_idx, 8, notes)
    ws.row_dimensions[row_idx].height = 65

# Summary row
summary_row = len(data) + 3
overall_be = sum(1 for d in data if d[0] and d[4].startswith("✅"))
overall_fe = sum(1 for d in data if d[0] and d[5].startswith("✅"))
overall_ok = sum(1 for d in data if d[0] and d[6].startswith("✅"))
overall_partial = sum(1 for d in data if d[0] and d[6].startswith("⚠️"))
overall_missing = sum(1 for d in data if d[0] and d[6].startswith("❌"))
total_ucs = overall_ok + overall_partial + overall_missing

ws.cell(row=summary_row, column=1, value="TOTAL USE CASES").font = Font(bold=True, size=12)
ws.cell(row=summary_row + 1, column=1, value=f"{total_ucs} total use cases across 17 modules")
ws.cell(row=summary_row + 2, column=1, value=f"✅ Fully Implemented: {overall_ok} ({overall_ok*100//total_ucs}%)") .font = green_font
ws.cell(row=summary_row + 3, column=1, value=f"⚠️ Partially Implemented: {overall_partial} ({overall_partial*100//total_ucs}%)").font = yellow_font
ws.cell(row=summary_row + 4, column=1, value=f"❌ Not Implemented: {overall_missing} ({overall_missing*100//total_ucs}%)").font = red_font
ws.cell(row=summary_row + 5, column=1, value=f"Backend endpoints done: {overall_be}/{total_ucs}")
ws.cell(row=summary_row + 5, column=2, value=f"Frontend pages done: {overall_fe}/{total_ucs}")

ws.auto_filter.ref = f"A1:H{len(data) + 1}"
ws.freeze_panes = "A2"

filepath = "D:\\Personal\\ACUITY-LMS-AI-Teacher\\BRD_Implementation_Status.xlsx"
wb.save(filepath)
print(f"Excel saved to {filepath}")
print(f"Total UCs: {total_ucs}")
print(f"  ✅ Fully Implemented: {overall_ok}")
print(f"  ⚠️  Partially Implemented: {overall_partial}")
print(f"  ❌ Not Implemented: {overall_missing}")
